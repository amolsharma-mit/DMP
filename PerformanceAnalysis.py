import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# -------------------------------------------------------------
# Helper: Calculate Precision, Recall, F1, Accuracy
# -------------------------------------------------------------
def compute_metrics(gs_set, dmp_set):
    tp = len(gs_set.intersection(dmp_set))
    fp = len(dmp_set - gs_set)
    fn = len(gs_set - dmp_set)
    tn = 0  # Not applicable in information extraction evaluation

    precision = tp / (tp + fp) if (tp + fp) else 0
    recall = tp / (tp + fn) if (tp + fn) else 0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0
    accuracy = tp / (tp + fp + fn) if (tp + fp + fn) else 0

    return {
        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
        "Precision": precision, "Recall": recall,
        "F1-Score": f1, "Accuracy": accuracy
    }


# -------------------------------------------------------------
# Helper: Apply color formatting (openpyxl)
# -------------------------------------------------------------
fill_match = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # light green
fill_missing = PatternFill(start_color="AFEEEE", end_color="AFEEEE", fill_type="solid") # aqua
fill_extra = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")   # light orange


# -------------------------------------------------------------
# Write results with colors to Excel
# -------------------------------------------------------------
def write_colored_rows(ws, data, is_relationship=False):
    """
    data = list of dicts:
      {"type": "match"/"missing"/"extra", "GS": ..., "DMP": ...}
    """

    # Write header
    if is_relationship:
        ws.append(["GS_subject", "GS_predicate", "GS_object",
                   "DMP_subject", "DMP_predicate", "DMP_object"])
    else:
        ws.append(["GS", "DMP"])

    # Write rows with colors
    for row in data:
        if is_relationship:
            gs = row["GS"]
            dmp = row["DMP"]

            gs_vals = list(gs) if gs else ["", "", ""]
            dmp_vals = list(dmp) if dmp else ["", "", ""]

            ws.append(gs_vals + dmp_vals)
            excel_row = ws.max_row

            # Color logic
            if row["type"] == "match":
                fill = fill_match
            elif row["type"] == "missing":
                fill = fill_missing
            else:
                fill = fill_extra

            for col in range(1, 7):
                ws.cell(row=excel_row, column=col).fill = fill

        else:
            ws.append([row["GS"], row["DMP"]])
            excel_row = ws.max_row

            if row["type"] == "match":
                fill = fill_match
            elif row["type"] == "missing":
                fill = fill_missing
            else:
                fill = fill_extra

            ws.cell(row=excel_row, column=1).fill = fill
            ws.cell(row=excel_row, column=2).fill = fill


# -------------------------------------------------------------
# CLASS PERFORMANCE
# -------------------------------------------------------------
def evaluate_classes(input_file, output_file):
    df_gs = pd.read_excel(input_file, sheet_name="GS_Classes", header=None)
    df_dmp = pd.read_excel(input_file, sheet_name="DMP_Classes", header=None)

    gs_set = set(
        df_gs[0]
        .dropna()
        .astype(str)
        .str.strip()
        .str.lower()
        .replace("", pd.NA)
        .dropna()
    )

    dmp_set = set(
        df_dmp[0]
        .dropna()
        .astype(str)
        .str.strip()
        .str.lower()
        .replace("", pd.NA)
        .dropna()
    )

    # Prepare table rows
    all_items = sorted(gs_set.union(dmp_set))
    rows = []

    for item in all_items:
        if item in gs_set and item in dmp_set:
            rows.append({"type": "match", "GS": item, "DMP": item})
        elif item in gs_set:
            rows.append({"type": "missing", "GS": item, "DMP": ""})
        else:
            rows.append({"type": "extra", "GS": "", "DMP": item})

    metrics = compute_metrics(gs_set, dmp_set)

    # Output Excel
    wb = Workbook()

    # Sheet 1: Data
    ws_data = wb.active
    ws_data.title = "Data"
    write_colored_rows(ws_data, rows, is_relationship=False)

    # Sheet 2: Metrics
    ws_m = wb.create_sheet("Performance Metrics")
    for k, v in metrics.items():
        ws_m.append([k, v])

    wb.save(output_file)


# -------------------------------------------------------------
# RELATIONSHIP PERFORMANCE
# -------------------------------------------------------------
def evaluate_relationships(input_file, output_file):
    df_gs = pd.read_excel(input_file, sheet_name="GS_Relationships")
    df_dmp = pd.read_excel(input_file, sheet_name="DMP_Relationships")

    # Convert each row to tuple (s,p,o)
    gs_set = set(
        tuple(str(v).strip().lower() for v in row)
        for row in df_gs.dropna().values
    )

    dmp_set = set(
        tuple(str(v).strip().lower() for v in row)
        for row in df_dmp.dropna().values
    )

    # Prepare sorted list by subject
    all_items = sorted(gs_set.union(dmp_set), key=lambda x: x[0])

    rows = []
    for rel in all_items:
        if rel in gs_set and rel in dmp_set:
            rows.append({"type": "match", "GS": rel, "DMP": rel})
        elif rel in gs_set:
            rows.append({"type": "missing", "GS": rel, "DMP": None})
        else:
            rows.append({"type": "extra", "GS": None, "DMP": rel})

    metrics = compute_metrics(gs_set, dmp_set)

    # Output Excel
    wb = Workbook()

    # Data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    write_colored_rows(ws_data, rows, is_relationship=True)

    # Metrics sheet
    ws_m = wb.create_sheet("Performance Metrics")
    for k, v in metrics.items():
        ws_m.append([k, v])

    wb.save(output_file)


# -------------------------------------------------------------
# MAIN WRAPPER
# -------------------------------------------------------------
if __name__ == "__main__":
    # Example usage:
    evaluate_classes("classes_input.xlsx", "classes_output.xlsx")
    evaluate_relationships("relationships_input.xlsx", "relationships_output.xlsx")

    print("Evaluation completed successfully.")
